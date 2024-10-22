import streamlit as st
import pandas as pd
from datetime import datetime, date
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

st.set_page_config(layout="wide")

# Custom CSS to increase font size and adjust other styles
st.markdown(
    """
    <style>
    label {
        font-size: 24px !important;
    }
    input, select, textarea, button {
        font-size: 24px !important;
    }
    h1 {
        font-size: 36px !important;
    }
    h2, h3, h4, h5, h6 {
        font-size: 32px !important;
    }
    .stTextInput, .stSelectbox {
        font-size: 24px !important;
    }
    .stButton>button {
        font-size: 24px !important;
    }
    .dataframe {
        font-size: 32px !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# File path for saving expense history
EXPENSE_HISTORY_FILE = "employee_expense_history_with_total.csv"

# Function to calculate fare
def calculate_fare(source, destination):
    fare = fare_data[(fare_data['Source'] == source) & (fare_data['Destination'] == destination)]['Price'].values
    if len(fare) > 0:
        return fare[0]
    else:
        return None

# Function to load existing expenses from the file
def load_expense_history():
    if os.path.exists(EXPENSE_HISTORY_FILE):
        return pd.read_csv(EXPENSE_HISTORY_FILE)
    else:
        return pd.DataFrame(columns=['Employee', 'Source', 'Destination', 'Fare', 'Date', 'Month'])

# Function to save the current session expenses to the file
def save_expense_history(expense_df):
    expense_df.to_csv(EXPENSE_HISTORY_FILE, index=False)

# Load fare data from Excel file
fare_data = pd.read_excel('fare_data.xlsx')

# Store employee expense history (in-memory and file-based)
if 'expense_history' not in st.session_state:
    st.session_state['expense_history'] = load_expense_history().to_dict('records')

# Function to generate Excel with multiple sheets and formatting
def generate_excel_with_formatting(expense_df):
    # Create a new workbook
    wb = Workbook()

    # Group the expense data by employee
    employees = expense_df['Employee'].unique()

    for employee in employees:
        # Create a new sheet for each employee
        ws = wb.create_sheet(title=employee)

        # Filter data for the employee
        employee_data = expense_df[expense_df['Employee'] == employee]

        # Sort data by date
        employee_data = employee_data.sort_values(by='Date')

        # Adding titles to the Excel
        ws['A1'] = employee
        ws['A1'].font = Font(bold=True, size=20)

        # Prepare the data for each month
        for month, month_data in employee_data.groupby('Month'):
            # Add month header
            ws.append([f'{month}'])
            month_row_start = ws.max_row  # Keep track of where the month's rows start

            # Append the data for that month
            for r in dataframe_to_rows(month_data, index=False, header=True):
                ws.append(r)

            # Style headers
            for col in range(1, 6):  # Assumes 5 columns in the table
                cell = ws.cell(row=month_row_start, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Add total row for that month
            total_fare = month_data['Fare'].sum()
            ws.append([None, None, None, 'Total:', f"${total_fare:.2f}"])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on cells that might be empty
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

    # Remove the default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Save the workbook to an in-memory file
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# Streamlit Layout: Two columns, left for instructions, right for form
left_column, right_column = st.columns([1, 2])

# Instructions in the left column
with left_column:
    st.markdown("## How to Use This App")
    st.write("""
        - **Enter Employee Name**: 开始输入员工的姓名。
        - **Select Source and Destination**: 选择源站和目的地站。
        - **Select Date**: 您可以选择具体的旅行日期或将其默认为今天的日期。
        - **Add Expense**: 选择所有内容后，单击“添加费用”以保存记录。
        - **View Expenses**: 您可以通过点击“查看费用”来查看旧记录 .
        - **Download Expenses**: 您可以通过点击“下载费用报告”来下载所有费用的报告。
    """)

# Form in the right column
with right_column:
    st.title("Employee Travel Expense Tracker")

    # Input for employee name
    employee_name = st.text_input("Enter Employee Name  (员工姓名)")

    # Ensure the first letter is uppercase and the rest lowercase
    employee_name = employee_name.capitalize()

    # Dropdowns for source and destination
    source_station = st.selectbox("Select Source Station(源站)", fare_data['Source'].unique())
    destination_station = st.selectbox("Select Destination Station(目的地站)", fare_data['Destination'].unique())

    # Date input (allows user to select a date or default to today's date)
    selected_date = st.date_input("Select Date (选择日期)", value=date.today())

    # Button to calculate fare and add to history
    if st.button("Add Expense (添加费用)"):
        if employee_name and source_station and destination_station:
            # Calculate fare
            fare = calculate_fare(source_station, destination_station)
            if fare is not None:
                # Add selected or current date
                date_str = selected_date.strftime("%Y-%m-%d")  # Date only (no time)
                month_str = selected_date.strftime("%B")  # Extract full month name (e.g., 'October')

                # Add to session and global expense history
                st.session_state['expense_history'].append({
                    'Employee': employee_name,  # Save the capitalized name
                    'Source': source_station,
                    'Destination': destination_station,
                    'Fare': fare,
                    'Date': date_str,
                    'Month': month_str
                })

                # Convert to DataFrame and save to file
                expense_df = pd.DataFrame(st.session_state['expense_history'])
                save_expense_history(expense_df)
                st.success(f"Fare added for {employee_name}: ${fare} on {date_str}")
            else:
                st.error("Fare not found for this route.")
        else:
            st.error("Please provide all required details.")

    # Button to view existing expense history
    if st.button("View Expenses(查看费用)"):
        st.subheader("View Expense History")

        # Load and group expenses by employee
        expense_df = load_expense_history()

        if not expense_df.empty:
            employees = expense_df['Employee'].unique()

            for employee in employees:
                st.write(f"### Expenses for {employee}")
                employee_expenses = expense_df[expense_df['Employee'] == employee]

                # Display the employee-specific expense table
                st.dataframe(employee_expenses, use_container_width=True)

                # Calculate and display total fare for each month
                st.write(f"#### Monthly Totals for {employee}:")
                monthly_totals = employee_expenses.groupby('Month')['Fare'].sum().reset_index()
                st.dataframe(monthly_totals)

                # Calculate and display the grand total for the employee
                grand_total = employee_expenses['Fare'].sum()
                st.write(f"**Grand Total Fare for {employee}:** ${grand_total:.2f}")
        else:
            st.write("No expense history found.")

    # Button to download the total expense report with formatting
    if st.button("Download Expense Report(下载格式化的费用报告)"):
        st.subheader("Download Expense Report with Formatting")

        # Load all expenses from file
        expense_df = load_expense_history()

        if not expense_df.empty:
            # Generate downloadable Excel data with formatting
            excel_data = generate_excel_with_formatting(expense_df)

            # Create the download button with the generated Excel data
            st.download_button(
                label="Download Total Expense Report as Excel",
                data=excel_data,
                file_name='total_expense_report.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            st.write("No expense history found to download.")
